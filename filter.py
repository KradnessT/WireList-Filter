from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("raw_data.xlsx") # 設定要執行的檔案名稱
ws = wb["QAQ"] # 設定要執行活頁簿名稱
row_count = 70 # 看WireList多寡設定作用到第幾行

def Panel_name(row_count):
    for col in range(1,2):
        for row in range(7,row_count):
            char = get_column_letter(col)
            cells_1 = ws[char + str(row)].value
            row1 = row + 1
            if cells_1 is None: 
                cells_1 = ""
            elif "<->" in str(cells_1):
                ws.merge_cells("A" + str(row) + ":E" + str(row)) # '盤名<->盤名'合併
                for row_1 in range(row1,row_count):
                    cells_2 = ws[char + str(row_1)].value
                    row_2 = row_1 - 1
                    if cells_1 is None:
                        cells_1 = ""
                    elif "<->" in str(cells_2):
                        ws.merge_cells("B" + str(row1) + ":B" + str(row_2)) # 'Symbol ID'合併
                        break

def Item(row_count):
    for col in range(1,2):
        for row in range(7,row_count):
            char = get_column_letter(col)
            cells_1 = ws[char + str(row)].value
            row1 = row + 1
            if cells_1 is None:
                cells_1 = ""
            elif "<->" not in str(cells_1) and str(cells_1) != "":
                for row_1 in range(row1,row_count):
                    cells_2 = ws[char + str(row_1)].value
                    row_2 = row_1 - 1
                    cells_3 = ws["I" + str(row)].value
                    if cells_2 is None:
                        cells_2 = ""
                    elif str(cells_2) != "":
                        ws.merge_cells("A" + str(row) + ":A" + str(row_2)) # 'Item'合併
                        ws.merge_cells("F" + str(row) + ":F" + str(row_2)) # 'Cable NO'合併
                        ws.merge_cells("G" + str(row) + ":G" + str(row_2)) # 'SPEC'合併
                        if str(cells_3) != "N/A":
                            ws.merge_cells("K" + str(row) + ":K" + str(row_2))
                            ws.merge_cells("L" + str(row) + ":L" + str(row_2))
                        elif str(cells_3) == "N/A":
                            ws.merge_cells("I" + str(row) + ":M" + str(row_2))
                        break

Panel_name(row_count)
Item(row_count)

wb.save("raw_data.xlsx")
